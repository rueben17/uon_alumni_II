# apps/home/management/commands/import_legacy_memberships.py
"""
Imports the legacy Google Forms membership register (docs/data/*.xlsx,
gitignored -- real member PII, never committed).

Built as a plain management command, not a django-import-export Resource/
ModelResource (2026-08-07): one row fans out into several model instances
(User, UserProfile, AlumniProfile, Membership, plus up to 2 overflow
AlumniQualification rows, 1 overflow AlumniEmploymentRecord, 1 overflow
AlumniPhoneNumber, 2 secondary EmailAddress rows) -- ModelResource assumes
one row = one model, and fighting that abstraction for this shape wasn't
worth it. Still reads the file the same way import-export would
(tablib-compatible), just without forcing the row-to-instance mapping
through its Resource class.

SAFE BY DEFAULT: runs inside a transaction that's rolled back unless
--commit is passed. Always run without --commit first and read the
report -- it lists every distinct value in the ambiguous columns
(Title/Gender/Nationality/CurrentCategory/College/Faculty) so gaps in the
mapping tables below surface before anything is written, not after.

Column formats confirmed against docs/data/Membership format.xlsx's 2
sample rows (2026-08-07) -- MOST of the mapping tables below only have the
handful of values seen so far. Re-run --dry-run against the real full
file and extend *_MAP with whatever UNMAPPED_VALUES reports before ever
using --commit for real.
"""
import json
import re
from datetime import date
from pathlib import Path

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.home.models import (
    AlumniEmploymentRecord,
    AlumniPhoneNumber,
    AlumniProfile,
    AlumniQualification,
    Faculty,
    Membership,
    MembershipTier,
    Qualification,
)
from apps.user.models import Gender, Honorific, UserProfile
from apps.user.phone import InvalidPhoneNumber, normalize_phone

User = get_user_model()

# ─────────────────────────────────────────────
# Mapping tables -- extend these once the real file's full value range
# is known (this command's dry-run report tells you exactly what's
# missing; don't guess ahead of that).
# ─────────────────────────────────────────────

HONORIFIC_MAP = {
    "MR": Honorific.MR, "MRS": Honorific.MRS, "MS": Honorific.MS,
    "MISS": Honorific.MISS, "DR": Honorific.DR, "PROF": Honorific.PROF,
    "REV": Honorific.REV, "HON": Honorific.HON, "ESQ": Honorific.ESQ,
}

GENDER_MAP = {
    "MALE": Gender.MALE, "FEMALE": Gender.FEMALE, "OTHER": Gender.OTHER,
    "M": Gender.MALE, "F": Gender.FEMALE,
}

NATIONALITY_MAP = {
    "KENYA": "Kenyan", "UGANDA": "Ugandan", "TANZANIA": "Tanzanian",
    "RWANDA": "Rwandan", "SOUTH SUDAN": "South Sudanese",
}

# Subscription -> Membership.Status. "Inactive" means EXPIRED (they were
# a real member at some point -- membership number assigned -- not PENDING,
# which implies never-yet-confirmed) -- Association decision 2026-08-07.
STATUS_MAP = {
    "ACTIVE": Membership.Status.ACTIVE,
    "INACTIVE": Membership.Status.EXPIRED,
}

# CurrentCategory -> MembershipTier.name (must match seeded names exactly,
# see apps/home/management/commands/seed_membership_tiers.py).
TIER_NAME_MAP = {
    "ANNUAL": "Full Annual Member",
    "STUDENT": "Student Annual Membership",
    "BRONZE": "Bronze Life Member",
    "SILVER": "Silver Life Member",
    "GOLD": "Gold Life Member",
    "PLATINUM": "Platinum Life Membership",
    "DIAMOND": "Diamond Life Membership",
    "CORPORATE": "Corporate Membership",
    "HONORARY": "Honorary Member",
}

TRUE_STRINGS = {"true", "yes", "y"}

FACULTY_MAPPING_PATH = Path(__file__).resolve().parents[4] / "docs" / "uon_faculty_mapping.json"

# Faculty code -> seeded Faculty.faculty_name (apps/staff/management/
# commands/seed_university_structure.py's exact strings).
FACULTY_CODE_TO_NAME = {
    "FAG": "Agriculture",
    "FVM": "Veterinary Medicine",
    "FEN": "Engineering",
    "FBED": "Built Environment and Design",
    "FST": "Science and Technology",
    "FED": "Education",
    "FHS": "Health Sciences",
    "FASS": "Arts and Social Sciences",
    "FLAW": "Law",
    "FBMS": "Business & Management Sciences",
}


class Command(BaseCommand):
    help = "Import the legacy Google Forms membership register (.xlsx/.csv)"

    def add_arguments(self, parser):
        parser.add_argument("file_path", help="Path to the .xlsx export")
        parser.add_argument(
            "--commit", action="store_true",
            help="Actually write. Without this, runs in a transaction that's always rolled back.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        self.faculty_mapping = json.loads(FACULTY_MAPPING_PATH.read_text(encoding="utf-8"))
        self._faculty_cache = {}
        self.unmapped = {
            "Title": set(), "Gender": set(), "Nationality": set(),
            "CurrentCategory": set(), "College": set(), "Faculty": set(),
            "Course": set(),
        }
        self.stats = {"created": 0, "updated": 0, "skipped_no_email": 0, "errors": 0}
        self.error_rows = []
        self.phone_conflicts = []

        rows = list(self._read_rows(file_path))
        self.stdout.write(f"Read {len(rows)} data row(s) from {file_path}")

        with transaction.atomic():
            for i, row in enumerate(rows, start=2):  # row 1 is the header
                # Each row gets its OWN nested atomic() (2026-09-04), NOT
                # the raw transaction.savepoint()/savepoint_rollback() API
                # this used to call directly. Those low-level calls don't
                # reset the connection's needs_rollback flag the way a
                # nested `with transaction.atomic():` does -- so on a
                # database-level error (e.g. IntegrityError) the raw form
                # left the connection marked broken, and EVERY row after
                # the first such failure then failed too, with a confusing
                # TransactionManagementError on the very next query
                # (confirmed: it broke even the final savepoint_rollback()
                # below, crashing the whole dry run instead of finishing
                # with a report). Nested atomic() is what Django's own
                # docs recommend for exactly this reason.
                try:
                    with transaction.atomic():
                        self._import_row(row, row_number=i)
                except Exception as exc:  # noqa: BLE001 -- one bad row must not abort the batch
                    self.stats["errors"] += 1
                    self.error_rows.append((i, str(exc)))

            if options["commit"]:
                self.stdout.write(self.style.WARNING("--commit passed: changes WRITTEN."))
            else:
                # set_rollback(True), not a raw savepoint call -- forces
                # THIS atomic() block to roll back on exit via Django's own
                # tracking, same reasoning as above.
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("Dry run (no --commit): all changes rolled back."))

        self._print_report()

    # ------------------------------------------------------------------
    # Reading
    # ------------------------------------------------------------------

    def _read_rows(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # First sheet with more than 1 row (the sample file has an
        # incidental empty "Sheet1" ahead of the real data sheet).
        ws = next(
            (wb[name] for name in wb.sheetnames if wb[name].max_row > 1),
            wb[wb.sheetnames[0]],
        )
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2):
            values = {h: c.value for h, c in zip(headers, row) if h}
            if any(v not in (None, "") for v in values.values()):
                yield values

    # ------------------------------------------------------------------
    # Per-row import -- the fan-out
    # ------------------------------------------------------------------

    def _import_row(self, row, row_number):
        def s(key):
            v = row.get(key)
            return str(v).strip() if v not in (None, "") else ""

        email = s("Email").lower()
        if not email:
            self.stats["skipped_no_email"] += 1
            self.error_rows.append((row_number, "No Email -- skipped (User.email is required+unique)"))
            return

        title = s("Title").upper()
        gender = s("Gender").upper()
        nationality_raw = s("Nationality").upper()
        if title and title not in HONORIFIC_MAP:
            self.unmapped["Title"].add(title)
        if gender and gender not in GENDER_MAP:
            self.unmapped["Gender"].add(gender)
        if nationality_raw and nationality_raw not in NATIONALITY_MAP:
            self.unmapped["Nationality"].add(nationality_raw)

        user, created = User.objects.get_or_create(email=email)
        self.stats["created" if created else "updated"] += 1

        phone = self._safe_phone(s("Telephone"))
        self._assign_phone_if_unclaimed(user, phone, row_number)
        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.honorific = HONORIFIC_MAP.get(title, profile.honorific)
        profile.given_name = s("FirstName").title() or profile.given_name
        profile.family_name = s("OtherNames").title() or profile.family_name
        profile.gender = GENDER_MAP.get(gender, profile.gender)
        profile.nationality = NATIONALITY_MAP.get(nationality_raw, s("Nationality").title() or profile.nationality or "Kenyan")
        profile.national_id = s("IdNo") or profile.national_id
        profile.postal_address = s("Address") or profile.postal_address
        profile.postal_code = s("PostalCode") or profile.postal_code
        profile.city = s("Town") or profile.city
        alt_phone = self._safe_phone(s("Telephone1"))
        if alt_phone:
            profile.alt_phone = alt_phone
        profile.save()

        # Overflow: Telephone2 -> AlumniPhoneNumber; Email1/Email2 -> allauth EmailAddress
        tel2 = self._safe_phone(s("Telephone2"))
        if tel2:
            AlumniPhoneNumber.objects.get_or_create(user=user, phone=tel2, defaults={"label": "legacy import"})
        for email_key in ("Email1", "Email2"):
            alt_email = s(email_key).lower()
            if alt_email:
                from allauth.account.models import EmailAddress
                EmailAddress.objects.get_or_create(
                    user=user, email=alt_email, defaults={"verified": False, "primary": False},
                )

        # AlumniProfile -- primary academic/employment slot
        alumni, _ = AlumniProfile.objects.get_or_create(user=user)
        primary_faculty = self._resolve_faculty(s("College"), s("Faculty"))
        if primary_faculty:
            alumni.faculty = primary_faculty
        grad_year = self._safe_year(s("Graduation"))
        if grad_year:
            # Legacy import data only ever has a year, never a real day/
            # month -- Jan 1 is the same placeholder the old
            # IntegerField-backed form widget used to reconstruct, kept
            # here for consistency with historical rows.
            alumni.graduation_date = date(grad_year, 1, 1)
        course_text = s("Course")
        alumni.qualification = self._resolve_qualification(course_text, primary_faculty)
        alumni.qualification_name_raw = course_text or alumni.qualification_name_raw
        alumni.current_employer = s("Organization") or alumni.current_employer
        alumni.employment_position = s("Position") or alumni.employment_position
        alumni.save()

        # Overflow qualifications: (College1,Faculty1,Course1,Graduation1) and (...2)
        for order, suffix in enumerate(("1", "2"), start=1):
            college = s(f"College{suffix}")
            faculty_text = s(f"Faculty{suffix}")
            course = s(f"Course{suffix}")
            grad = s(f"Graduation{suffix}")
            if not any([college, faculty_text, course, grad]):
                continue
            overflow_faculty = self._resolve_faculty(college, faculty_text)
            AlumniQualification.objects.get_or_create(
                alumni_profile=alumni, order=order,
                defaults={
                    "legacy_college": college,
                    "faculty": overflow_faculty,
                    "qualification": self._resolve_qualification(course, overflow_faculty),
                    "course_name_raw": course,
                    "graduation_year": self._safe_year(grad),
                },
            )

        # Overflow employment: (Position1, Organization1)
        position1 = s("Position1")
        organization1 = s("Organization1")
        if position1 or organization1:
            AlumniEmploymentRecord.objects.get_or_create(
                alumni_profile=alumni, order=1,
                defaults={"organization": organization1, "position": position1},
            )

        # Membership -- set fields directly rather than through activate()/
        # record_installment_payment(): those are for NEW forward-looking
        # transactions, not backfilling a historical row against today's
        # tier settings.
        category = s("CurrentCategory").upper()
        if category and category not in TIER_NAME_MAP:
            self.unmapped["CurrentCategory"].add(category)
        tier_name = TIER_NAME_MAP.get(category)
        tier = MembershipTier.objects.filter(name=tier_name).first() if tier_name else None
        if tier is None:
            self.error_rows.append((row_number, f"Unresolved CurrentCategory {category!r} -- membership not created"))
            return

        mem_id = s("MemId")
        status_raw = s("Subscription").upper()
        started_on = row.get("JoinDate")
        started_on = started_on.date() if hasattr(started_on, "date") else None

        membership, mem_created = Membership.objects.get_or_create(
            user=user, tier=tier,
            defaults={
                "membership_number": mem_id or None,
                "status": STATUS_MAP.get(status_raw, Membership.Status.EXPIRED),
                "is_lifetime": tier.is_lifetime(),
                "started_on": started_on,
                "card_issued": s("Card").upper() in TRUE_STRINGS,
                "legacy_signed": s("Signed").lower() in TRUE_STRINGS,
            },
        )
        if not mem_created:
            return  # already imported this user+tier combo -- idempotent, leave it alone

    # ------------------------------------------------------------------
    # Resolvers
    # ------------------------------------------------------------------

    def _resolve_faculty(self, college_code, faculty_text):
        college_code = college_code.upper()
        faculty_text = faculty_text.upper()
        if not college_code and not faculty_text:
            return None

        cache_key = (college_code, faculty_text)
        if cache_key in self._faculty_cache:
            return self._faculty_cache[cache_key]

        result = None

        # 1. College code alone, if unambiguous (single successor faculty).
        for legacy in self.faculty_mapping["legacy_colleges"]:
            if legacy["code"] == college_code and len(legacy["successor_faculties"]) == 1:
                name = FACULTY_CODE_TO_NAME.get(legacy["successor_faculties"][0])
                result = Faculty.objects.filter(faculty_name=name).first()
                break

        # 2. Legacy unit/faculty text, fuzzy-matched against unit_mapping.
        if result is None and faculty_text:
            for unit in self.faculty_mapping["unit_mapping"]:
                legacy_unit = unit["legacy_unit"].upper()
                if faculty_text in legacy_unit or legacy_unit in faculty_text:
                    name = FACULTY_CODE_TO_NAME.get(unit["current_faculty"])
                    if name:
                        result = Faculty.objects.filter(faculty_name=name).first()
                        break

        # 3. Faculty text matches a CURRENT faculty name directly.
        if result is None and faculty_text:
            result = Faculty.objects.filter(faculty_name__iexact=faculty_text).first()

        if result is None:
            if college_code:
                self.unmapped["College"].add(college_code)
            if faculty_text:
                self.unmapped["Faculty"].add(faculty_text)

        self._faculty_cache[cache_key] = result
        return result

    def _resolve_qualification(self, course_text, faculty):
        if not course_text:
            return None
        qs = Qualification.objects.all()
        if faculty:
            qs = qs.filter(faculty=faculty)
        match = qs.filter(name__iexact=course_text).first() or qs.filter(name__icontains=course_text).first()
        if match is None:
            self.unmapped["Course"].add(course_text)
        return match

    def _assign_phone_if_unclaimed(self, user, phone, row_number):
        """User.phone is unique=True, but the legacy register has no such
        guarantee -- two different real people can share an office
        landline (2026-09-04). A member's own mobile is what they'd
        actually search on to find their profile, not this historical
        field, so treat it as best-effort: skip and log the conflict
        rather than fail the whole row over it. Checked proactively
        (not left to the database) so a collision reads as one line in
        the report, not an IntegrityError.
        """
        if not phone:
            return
        holder = User.objects.filter(phone=phone).exclude(pk=user.pk).first()
        if holder:
            self.phone_conflicts.append((row_number, phone, holder.email))
            return
        user.phone = phone

    def _safe_phone(self, value):
        if not value:
            return None
        try:
            return normalize_phone(value)
        except InvalidPhoneNumber:
            return None

    def _safe_year(self, value):
        if not value:
            return None
        match = re.search(r"\d{4}", str(value))
        return int(match.group()) if match else None

    # ------------------------------------------------------------------
    # Report
    # ------------------------------------------------------------------

    def _print_report(self):
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== Import report ==="))
        for key, count in self.stats.items():
            self.stdout.write(f"  {key}: {count}")

        if self.error_rows:
            self.stdout.write(self.style.WARNING(f"\n{len(self.error_rows)} row-level issue(s):"))
            for row_number, msg in self.error_rows:
                self.stdout.write(f"  row {row_number}: {msg}")

        if self.phone_conflicts:
            self.stdout.write(self.style.WARNING(
                f"\n{len(self.phone_conflicts)} phone number(s) already claimed by another "
                "account -- left blank on the row below rather than failing it. Likely a "
                "shared office line; resolve by hand if it's actually the same person:"
            ))
            for row_number, phone, holder_email in self.phone_conflicts:
                self.stdout.write(f"  row {row_number}: {phone} already on {holder_email}")

        any_unmapped = any(self.unmapped.values())
        if any_unmapped:
            self.stdout.write(self.style.WARNING("\nUnmapped values found -- extend the *_MAP tables at the top of this file:"))
            for column, values in self.unmapped.items():
                if values:
                    self.stdout.write(f"  {column}: {sorted(values)}")
        else:
            self.stdout.write(self.style.SUCCESS("\nNo unmapped values -- every row's ambiguous fields resolved cleanly."))
