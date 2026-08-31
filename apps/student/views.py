from allauth.account.adapter import get_adapter
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.generic import CreateView, TemplateView, View
from django_q.tasks import async_task
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from apps.student import analytics
from apps.student.forms import SCORE_FIELDS, InterviewScoreSheetForm, StudentRegisterForm, score_field_max
from apps.student.models import County, Gender, ParentalStatus, ScholarshipApplication, Student
from apps.user.mixins import StaffOrSuperuserRequiredMixin

# Counties past this rank in the distribution collapse into a single
# "Other" bucket on the dashboard's bar chart -- a 47-bar chart is as
# unreadable as a 47-slice pie.
COUNTY_TOP_N = 15


@login_required
def all_uon_students(request):
    context = {}
    return render(request, "student/all_uon_students.html", context)


class StudentRegisterView(LoginRequiredMixin, CreateView):
    """
    One-time student sign-up. apps/user/adapter.py's login/signup
    redirect logic sends every authenticated user without a Student
    record here -- see its "Students: no auto-created stub" note for why
    this is a real form (mirroring AlumniRegisterView) rather than
    staff's auto-stub-then-complete flow.

    The @students.uonbi.ac.ke restriction is re-checked here, not just
    at the Google-login step in apps/user/adapter.py's pre_social_login():
    this view is also reachable via the bare domain's /students/ prefix
    (main/urls.py, kept for backward compatibility), which has no
    subdomain of its own for that check to key off. Re-checking here
    means the restriction holds regardless of which host served the
    page, not just the one true entry point.
    """
    model = Student
    form_class = StudentRegisterForm
    template_name = "student/register.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if hasattr(request.user, "student"):
                return redirect(self._post_register_url())

            from apps.user.adapter import ALLOWED_STUDENT_LOGIN_DOMAINS, RESTRICT_GOOGLE_LOGIN_DOMAINS

            if RESTRICT_GOOGLE_LOGIN_DOMAINS:
                domain = request.user.email.split("@")[-1].lower()
                if domain not in ALLOWED_STUDENT_LOGIN_DOMAINS:
                    messages.error(
                        request,
                        "Student sign-up needs your @students.uonbi.ac.ke Google account -- "
                        "please sign out and sign in again from the students portal.",
                    )
                    return redirect("home:uon_alumni_home")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.user = self.request.user
        # The students subdomain only accepts @students.uonbi.ac.ke Google
        # accounts (dispatch() above, apps/user/adapter.py's
        # pre_social_login()) -- sign-in IS this address, so there's
        # nothing further to ask for it.
        form.instance.student_email = self.request.user.email
        response = super().form_valid(form)
        messages.success(self.request, "You're signed up! You can now apply for the scholarship.")
        return response

    def get_success_url(self):
        return self._post_register_url()

    def _post_register_url(self):
        # Set by apps.home.views.uon_alumni_scholarship for an anonymous
        # or not-yet-registered visitor -- honoured here so finishing
        # sign-up lands them back where they started, not just on the
        # generic students landing page.
        next_url = self.request.session.pop("post_login_next", None)
        if next_url and get_adapter().is_safe_url(next_url):
            return next_url
        return reverse("all_uon_students")


class EvaluateApplicationView(StaffOrSuperuserRequiredMixin, View):
    """
    Two-pane interview scoring screen: the applicant's uploaded PDF
    alongside InterviewScoreSheet's scoring matrix. One score sheet per
    ScholarshipApplication (OneToOne, related_name="score_sheet") --
    get-or-create semantics, not plain create, so revisiting an already-
    scored application reopens it for editing instead of erroring.

    evaluator is never a form field (InterviewScoreSheetForm excludes
    both application and evaluator) -- both are set here from the URL
    and request.user.employee respectively, never from POST data.

    pk is optional (2026-08-14, student:evaluate_application_list route)
    -- with no pk, this renders just the applicant picker with nothing
    selected yet, the landing page the new navbar link and the picker's
    own "no selection" state both need. The picker itself is always in
    context/rendered regardless, letting staff jump to a different
    applicant from the two-pane screen too.
    """
    template_name = "student/evaluate_application.html"

    def _get_application(self, pk):
        if pk is None:
            return None
        return get_object_or_404(
            ScholarshipApplication.objects.select_related("student", "faculty", "department", "score_sheet"),
            pk=pk,
        )

    def _total_possible(self):
        return sum(score_field_max(name) for name in SCORE_FIELDS)

    def _context(self, application, form):
        # Bound fields, not raw names -- {{ form.field_name }} can't
        # resolve a dynamic template variable (Django's dotted-variable
        # lookup only splits literal tokens), so the per-criterion
        # {field, max} pairs are built here in Python instead, where
        # form[name] works with a variable just fine.
        score_rows = [(form[name], score_field_max(name)) for name in SCORE_FIELDS] if form else []
        # Lightweight, single query -- only the fields the picker's
        # <option> labels need, not full application rows.
        applicants = ScholarshipApplication.objects.only(
            "id", "first_name", "surname", "registration_number"
        ).order_by("surname", "first_name")
        return {
            "application": application,
            "form": form,
            "score_rows": score_rows,
            "total_possible": self._total_possible(),
            "applicants": applicants,
        }

    def get(self, request, pk=None):
        application = self._get_application(pk)
        form = None
        if application is not None:
            score_sheet = getattr(application, "score_sheet", None)
            initial = {} if score_sheet else {"course_undertaking": application.current_course}
            form = InterviewScoreSheetForm(instance=score_sheet, initial=initial)
        return render(request, self.template_name, self._context(application, form))

    def post(self, request, pk=None):
        application = self._get_application(pk)
        if application is None:
            # Nothing to score without an applicant -- the picker's own
            # navigation (GET) is how one gets chosen; POSTing here with
            # no pk isn't a reachable path through the rendered form.
            return redirect("evaluate_application_list")

        score_sheet = getattr(application, "score_sheet", None)
        form = InterviewScoreSheetForm(request.POST, instance=score_sheet)

        if form.is_valid():
            evaluator = getattr(request.user, "employee", None)
            if evaluator is None:
                form.add_error(None, "Your account has no staff Employee record -- cannot record an evaluation.")
            else:
                sheet = form.save(commit=False)
                sheet.application = application
                sheet.evaluator = evaluator
                sheet.save()
                sheet_id = sheet.pk
                transaction.on_commit(
                    lambda: async_task("apps.student.tasks.send_evaluation_receipt", sheet_id)
                )
                messages.success(request, f"Evaluation saved -- total score {sheet.total_score}/{self._total_possible()}.")
                return redirect("evaluate_application", pk=pk)

        return render(request, self.template_name, self._context(application, form))


class ApplicantDashboardView(StaffOrSuperuserRequiredMixin, TemplateView):
    """
    Applicant distribution by faculty/gender/county -- three charts,
    each backed by exactly one annotated .values().annotate(Count(...))
    query (no per-applicant Python loop). Renders fine on an empty
    ScholarshipApplication table -- .values().annotate() over zero rows
    is just an empty list, and Chart.js draws an empty axis/canvas for
    empty labels/data rather than erroring.
    """
    template_name = "student/applicant_dashboard.html"

    def _faculty_chart_data(self):
        rows = (
            ScholarshipApplication.objects.values("faculty__faculty_name")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        return {
            "labels": [row["faculty__faculty_name"] or "Unspecified" for row in rows],
            "counts": [row["count"] for row in rows],
        }

    def _gender_chart_data(self):
        rows = ScholarshipApplication.objects.values("gender").annotate(count=Count("id")).order_by("-count")
        labels = dict(Gender.choices)
        return {
            "labels": [labels.get(row["gender"], row["gender"]) for row in rows],
            "counts": [row["count"] for row in rows],
        }

    def _county_chart_data(self):
        rows = list(
            ScholarshipApplication.objects.values("county_of_residence")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        labels = dict(County.choices)
        # Splitting top-N from the rest below is Python work over the
        # already-aggregated result (at most 47 rows), not a loop over
        # ScholarshipApplication rows themselves -- the query above is
        # still the only trip to the database this chart makes.
        top, rest = rows[:COUNTY_TOP_N], rows[COUNTY_TOP_N:]
        chart_labels = [labels.get(row["county_of_residence"], row["county_of_residence"]) for row in top]
        chart_counts = [row["count"] for row in top]
        if rest:
            chart_labels.append("Other")
            chart_counts.append(sum(row["count"] for row in rest))
        return {"labels": chart_labels, "counts": chart_counts}

    def _pipeline_chart_data(self):
        totals = ScholarshipApplication.objects.aggregate(total=Count("id"), evaluated=Count("score_sheet"))
        pending = totals["total"] - totals["evaluated"]
        return {
            "labels": ["Evaluated", "Pending Evaluation"],
            "counts": [totals["evaluated"], pending],
        }

    def _parental_status_chart_data(self):
        rows = (
            ScholarshipApplication.objects.exclude(score_sheet__isnull=True)
            .values("score_sheet__parental_status")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        labels = dict(ParentalStatus.choices)
        return {
            "labels": [labels.get(row["score_sheet__parental_status"], row["score_sheet__parental_status"]) for row in rows],
            "counts": [row["count"] for row in rows],
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["faculty_data"] = self._faculty_chart_data()
        context["gender_data"] = self._gender_chart_data()
        context["county_data"] = self._county_chart_data()
        context["pipeline_data"] = self._pipeline_chart_data()
        context["parental_status_data"] = self._parental_status_chart_data()
        return context


# ---------------------------------------------------------------------
# Scholarship analytics Excel export -- apps/student/analytics.py holds
# the queries, everything below just renders their output into an
# .xlsx workbook. Kept as module-level functions (one per sheet), not
# methods, since none of them need view state.
# ---------------------------------------------------------------------

def _write_header_row(ws, headers, row=1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _autosize_columns(ws, headers, min_width=10, max_width=40):
    for col_idx, header in enumerate(headers, start=1):
        width = max(min_width, min(max_width, len(str(header)) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _round_or_zero(value, digits=2):
    return round(value, digits) if value is not None else 0


def _build_summary_sheet(ws, totals, distribution):
    """Metric groups 1 and 2 (exercise_totals, score_distribution) as
    label/value rows, plus the score-band bar chart."""
    headers = ["Metric", "Value"]
    _write_header_row(ws, headers)

    rows_data = [
        ("Total Applicants", totals["applicants"]),
        ("Evaluated", totals["evaluated"]),
        ("Unevaluated (Pending)", totals["unevaluated"]),
        ("Completion Rate (%)", _round_or_zero(totals["completion_rate"])),
        ("Distinct Evaluators", totals["distinct_evaluators"]),
        ("Earliest Evaluation Date", totals["earliest_evaluation"]),
        ("Latest Evaluation Date", totals["latest_evaluation"]),
        ("", ""),
        (
            "Note",
            "Score statistics below, and on every other sheet, cover EVALUATED "
            "applicants only. Unevaluated applicants are counted above but "
            "excluded from every score statistic.",
        ),
        ("", ""),
        ("Min Score", distribution["min"]),
        ("Max Score", distribution["max"]),
        ("Mean Score", _round_or_zero(distribution["mean"])),
        ("Median Score", _round_or_zero(distribution["median"])),
        ("Standard Deviation", _round_or_zero(distribution["stdev"])),
        ("Q1 (25th percentile)", _round_or_zero(distribution["q1"])),
        ("Q3 (75th percentile)", _round_or_zero(distribution["q3"])),
    ]
    for row_idx, (label, value) in enumerate(rows_data, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)

    _autosize_columns(ws, headers, min_width=22)
    ws.column_dimensions["B"].width = 60

    # Score-band table, placed clear of the label/value rows above, with
    # its own bar chart -- the one native chart this sheet needs.
    band_start_row = len(rows_data) + 4
    ws.cell(row=band_start_row, column=1, value="Score Band").font = Font(bold=True)
    ws.cell(row=band_start_row, column=2, value="Count").font = Font(bold=True)
    for i, band in enumerate(distribution["bands"], start=1):
        ws.cell(row=band_start_row + i, column=1, value=band["band"])
        ws.cell(row=band_start_row + i, column=2, value=band["count"])

    last_band_row = band_start_row + len(distribution["bands"])
    chart = BarChart()
    chart.title = "Score Distribution (bands)"
    chart.y_axis.title = "Applicants"
    data = Reference(ws, min_col=2, min_row=band_start_row, max_row=last_band_row)
    cats = Reference(ws, min_col=1, min_row=band_start_row + 1, max_row=last_band_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")


def _build_applicant_scores_sheet(ws, rows):
    """One row per evaluated applicant -- metric group data feeding
    this sheet is analytics.applicant_scores(), not one of the 8 metric
    groups (it's the underlying per-row data, not an aggregate)."""
    criterion_labels = [name.replace("score_", "").replace("_", " ").title() for name in SCORE_FIELDS]
    headers = (
        ["Registration Number", "Name", "Faculty", "Gender", "County"]
        + criterion_labels
        + ["Total Score", "Percentage of Max (%)", "Evaluator", "Evaluation Date"]
    )
    _write_header_row(ws, headers)

    for row_idx, row in enumerate(rows, start=2):
        col = 1
        for value in (row["registration_number"], row["name"], row["faculty_name"], row["gender_label"], row["county_label"]):
            ws.cell(row=row_idx, column=col, value=value)
            col += 1
        for name in SCORE_FIELDS:
            ws.cell(row=row_idx, column=col, value=row["scores"][name])
            col += 1
        ws.cell(row=row_idx, column=col, value=row["total"])
        col += 1
        ws.cell(row=row_idx, column=col, value=_round_or_zero(row["percentage_of_max"]))
        col += 1
        ws.cell(row=row_idx, column=col, value=row["evaluator_name"])
        col += 1
        ws.cell(row=row_idx, column=col, value=row["evaluation_date"])

    _autosize_columns(ws, headers)


def _build_by_faculty_sheet(ws, rows):
    headers = ["Faculty", "Count", "Share of Total (%)", "Mean Score", "Median Score", "Highest Score"]
    _write_header_row(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["faculty_name"])
        ws.cell(row=row_idx, column=2, value=row["count"])
        ws.cell(row=row_idx, column=3, value=_round_or_zero(row["share_of_total"]))
        ws.cell(row=row_idx, column=4, value=_round_or_zero(row["mean"]))
        ws.cell(row=row_idx, column=5, value=_round_or_zero(row["median"]))
        ws.cell(row=row_idx, column=6, value=row["highest"])
    _autosize_columns(ws, headers)

    if rows:
        last_row = len(rows) + 1
        chart = BarChart()
        chart.title = "Applicants Evaluated by Faculty"
        chart.y_axis.title = "Count"
        data = Reference(ws, min_col=2, min_row=1, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "H2")


def _build_by_county_sheet(ws, rows):
    headers = ["County", "Count", "Mean Score"]
    _write_header_row(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["county_label"])
        ws.cell(row=row_idx, column=2, value=row["count"])
        ws.cell(row=row_idx, column=3, value=_round_or_zero(row["mean"]))
    _autosize_columns(ws, headers)

    if rows:
        last_row = len(rows) + 1
        chart = BarChart()
        chart.title = "Applicants Evaluated by County"
        chart.y_axis.title = "Count"
        data = Reference(ws, min_col=2, min_row=1, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")


def _build_by_gender_sheet(ws, rows):
    headers = ["Gender", "Count", "Share (%)", "Mean Score", "Median Score"]
    _write_header_row(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["gender_label"])
        ws.cell(row=row_idx, column=2, value=row["count"])
        ws.cell(row=row_idx, column=3, value=_round_or_zero(row["share"]))
        ws.cell(row=row_idx, column=4, value=_round_or_zero(row["mean"]))
        ws.cell(row=row_idx, column=5, value=_round_or_zero(row["median"]))
    _autosize_columns(ws, headers)

    if rows:
        last_row = len(rows) + 1
        chart = PieChart()
        chart.title = "Applicants Evaluated by Gender"
        data = Reference(ws, min_col=2, min_row=1, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "G2")


def _build_criterion_diagnostics_sheet(ws, rows):
    headers = ["Criterion", "Mean", "Max Possible", "Mean (% of Max)", "Std Dev", "Low Variance Flag"]
    _write_header_row(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["label"])
        ws.cell(row=row_idx, column=2, value=_round_or_zero(row["mean"]))
        ws.cell(row=row_idx, column=3, value=row["max_possible"])
        ws.cell(row=row_idx, column=4, value=_round_or_zero(row["mean_pct_of_max"]))
        ws.cell(row=row_idx, column=5, value=_round_or_zero(row["stdev"]))
        ws.cell(row=row_idx, column=6, value="Yes" if row["low_variance_flag"] else "No")
    _autosize_columns(ws, headers)


def _build_evaluator_consistency_sheet(ws, rows):
    headers = ["Evaluator", "Applicants Evaluated", "Mean Score Awarded", "Difference from Overall Mean"]
    _write_header_row(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["evaluator_name"])
        ws.cell(row=row_idx, column=2, value=row["count"])
        ws.cell(row=row_idx, column=3, value=_round_or_zero(row["mean"]))
        ws.cell(row=row_idx, column=4, value=_round_or_zero(row["difference_from_overall_mean"]))
    _autosize_columns(ws, headers)


def _build_cutoff_simulation_sheet(ws, rows):
    if not rows:
        headers = ["Decile (%)", "Threshold", "Surviving Count"]
        _write_header_row(ws, headers)
        _autosize_columns(ws, headers)
        return

    gender_keys = list(rows[0]["by_gender"].keys())
    faculty_keys = list(rows[0]["by_faculty"].keys())
    headers = (
        ["Decile (%)", "Threshold", "Surviving Count"]
        + [f"{g} (survivors)" for g in gender_keys]
        + [f"{f} (survivors)" for f in faculty_keys]
    )
    _write_header_row(ws, headers)

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["decile"])
        ws.cell(row=row_idx, column=2, value=row["threshold"])
        ws.cell(row=row_idx, column=3, value=row["surviving_count"])
        col = 4
        for gender in gender_keys:
            ws.cell(row=row_idx, column=col, value=row["by_gender"][gender])
            col += 1
        for faculty in faculty_keys:
            ws.cell(row=row_idx, column=col, value=row["by_faculty"][faculty])
            col += 1
    _autosize_columns(ws, headers)


class ScholarshipAnalyticsExportView(StaffOrSuperuserRequiredMixin, View):
    """Full scholarship-exercise analytics as a downloadable .xlsx
    workbook -- one sheet per apps/student/analytics.py metric group,
    plus a per-applicant Applicant Scores sheet. See that module's
    docstring for the schema ground truth this was built against."""

    def get(self, request):
        totals = analytics.exercise_totals()
        distribution = analytics.score_distribution()
        faculty_rows = analytics.by_faculty()
        county_rows = analytics.by_county()
        gender_rows = analytics.by_gender()
        criterion_rows = analytics.criterion_diagnostics()
        evaluator_rows = analytics.evaluator_consistency()
        cutoff_rows = analytics.cutoff_simulation()
        applicant_rows = analytics.applicant_scores()

        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "Summary"
        _build_summary_sheet(summary_ws, totals, distribution)

        _build_applicant_scores_sheet(wb.create_sheet("Applicant Scores"), applicant_rows)
        _build_by_faculty_sheet(wb.create_sheet("By Faculty"), faculty_rows)
        _build_by_county_sheet(wb.create_sheet("By County"), county_rows)
        _build_by_gender_sheet(wb.create_sheet("By Gender"), gender_rows)
        _build_criterion_diagnostics_sheet(wb.create_sheet("Criterion Diagnostics"), criterion_rows)
        _build_evaluator_consistency_sheet(wb.create_sheet("Evaluator Consistency"), evaluator_rows)
        _build_cutoff_simulation_sheet(wb.create_sheet("Cutoff Simulation"), cutoff_rows)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"scholarship_analytics_{timezone.now():%Y%m%d_%H%M}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
